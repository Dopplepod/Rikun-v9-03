from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET
from flask import Flask, jsonify, render_template, request, send_file

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
HIERARCHY_PATH = DATA_DIR / 'hierarchy.xml'
MAPPING_WORKBOOK_PATH = DATA_DIR / 'v9_mapping.xlsx'

app = Flask(__name__)

SUMMARY_ROWS = [
    ('4100000', 'Revenue'),
    ('EXP_EXC_DNA', 'Operating Expense (Ex-D And A)'),
    ('EBITDA', 'EBITDA'),
    ('EBIT', 'EBIT'),
    ('PBT', 'Profit Or Loss Before Tax'),
    ('PAT', 'Net Profit Or Loss (PAT)'),
    ('ATT_OWN', 'Profit Or Loss Attributable To Owners Of The Company'),
]

DRILLDOWN_ROWS = [
    ('4100000', 'Revenue'),
    ('EXP_EXC_DNA', 'Operating Expense (Ex-D And A)'),
    ('5010000', 'Cost Of Raw Materials And Supplies'),
    ('5020000', 'Staff Costs'),
    ('5030000', 'Licence Fees'),
    ('5050000', 'Company Premise Utilities And Maintenance'),
    ('5060000', 'Subcontracting services'),
    ('5080000', 'Other costs'),
    ('EBITDA', 'EBITDA'),
    ('5040000', 'Depreciation And Amortisation'),
    ('EBIT', 'EBIT'),
    ('6021000', 'Finance Income'),
    ('6031000', 'Finance Expense'),
    ('8010000', 'Share Of Results Of AJV'),
    ('6010000', 'Non operating gain loss'),
    ('6990000', 'Exceptional Items'),
    ('PBT', 'Profit Or Loss Before Tax'),
    ('6070000', 'Income Tax Expense'),
    ('8610000', 'Profit Or Loss From Discontinued Operation (Net Of Tax)'),
    ('PAT', 'Net Profit Or Loss (PAT)'),
    ('PL_MI', 'Minority Interest'),
    ('ATT_OWN', 'Profit Or Loss Attributable To Owners Of The Company'),
]

HIGHLIGHT_CODES = {'4100000', 'EXP_EXC_DNA', 'EBITDA', 'EBIT', 'PBT', 'PAT', 'ATT_OWN'}


@dataclass
class Node:
    code: str
    name: str
    parent: str | None
    level: int
    children: list[str] = field(default_factory=list)


def _stream_xlsx_rows(file_storage, header_row: int):
    stream = getattr(file_storage, 'stream', file_storage)
    stream.seek(0)
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == header_row:
            header = [str(x).strip() if x is not None else '' for x in row]
            continue
        if header is not None:
            yield header, row


def _build_index(header: list[str], required: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    missing: list[str] = []
    for name in required:
        if name not in header:
            missing.append(name)
        else:
            idx[name] = header.index(name)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return idx


def normalize_code(value: Any) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().upper().replace('–', '-').replace('—', '-').replace(' ', '').replace(',', '')
    if s.endswith('.0'):
        s = s[:-2]
    return '' if s == 'NAN' else s


def extract_gl_code(value: Any) -> str:
    s = normalize_code(value)
    if not s:
        return ''
    m = re.match(r'^([A-Z0-9_]+)', s)
    return m.group(1) if m else s


def first_nonblank(series: pd.Series) -> str:
    for v in series:
        if pd.notna(v):
            s = str(v).strip()
            if s and s.upper() != 'NAN':
                return s
    return ''


def parse_hierarchy(path: Path) -> dict[str, Node]:
    tree = ET.parse(path)
    root = tree.getroot()
    nodes: dict[str, Node] = {}

    def walk(elem: ET.Element, parent: str | None, level: int):
        code = normalize_code(elem.attrib.get('code'))
        name = (elem.attrib.get('name') or '').strip()
        if not code:
            return
        nodes[code] = Node(code=code, name=name, parent=parent, level=level)
        if parent and parent in nodes:
            nodes[parent].children.append(code)
        for child in elem.findall('./Account'):
            walk(child, code, level + 1)

    for acct in root.findall('./Account'):
        walk(acct, None, 0)
    return nodes


def ancestors_of(code: str, nodes: dict[str, Node]) -> set[str]:
    out: set[str] = set()
    cur = code
    seen = set()
    while cur and cur not in seen and cur in nodes:
        out.add(cur)
        seen.add(cur)
        cur = nodes[cur].parent or ''
    return out


def load_v7_mapping(path: Path, nodes: dict[str, Node]) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    os_df = pd.read_excel(path, sheet_name='ERP TO OS')
    os_df.columns = [str(c).strip() for c in os_df.columns]
    if 'Local COA' not in os_df.columns or 'OS Account' not in os_df.columns:
        raise ValueError('ERP TO OS sheet must contain Local COA and OS Account columns.')
    os_df = os_df.copy()
    os_df['gl_code'] = os_df['Local COA'].map(extract_gl_code)
    os_df['os_leaf_code'] = os_df['OS Account'].map(normalize_code)
    desc_col = next((c for c in ['Description', 'GL Name', 'Account name', 'Unnamed: 2'] if c in os_df.columns), None)
    if desc_col:
        os_df['os_description'] = os_df[desc_col].fillna('').astype(str).str.strip()
    else:
        os_df['os_description'] = ''
    os_df = os_df[(os_df['gl_code'] != '') & (os_df['os_leaf_code'] != '')].copy()

    bfc_raw = pd.read_excel(path, sheet_name='ERP to BFC', header=None)
    if bfc_raw.empty:
        raise ValueError('ERP to BFC sheet is empty.')

    header_row_idx = None
    for i in range(len(bfc_raw)):
        row_vals = [str(x).strip() if pd.notna(x) else '' for x in bfc_raw.iloc[i].tolist()]
        if 'GL Code' in row_vals and 'SAP Mapping' in row_vals:
            header_row_idx = i
            header = row_vals
            break
    if header_row_idx is None:
        raise ValueError('ERP to BFC sheet must contain GL Code and SAP Mapping columns.')

    bfc = bfc_raw.iloc[header_row_idx + 1:].copy()
    bfc.columns = header
    required = ['GL Code', 'SAP Mapping']
    missing = [c for c in required if c not in bfc.columns]
    if missing:
        raise ValueError(f"ERP to BFC sheet is missing required columns: {', '.join(missing)}")
    bfc['gl_code'] = bfc['GL Code'].map(extract_gl_code)
    bfc['sap_mapping'] = bfc['SAP Mapping'].map(normalize_code)
    if 'GL Name' in bfc.columns:
        bfc['gl_name'] = bfc['GL Name'].fillna('').astype(str).str.strip()
    else:
        bfc['gl_name'] = ''
    if 'SAP Description' in bfc.columns:
        bfc['sap_description'] = bfc['SAP Description'].fillna('').astype(str).str.strip()
    else:
        bfc['sap_description'] = ''
    bfc = bfc[bfc['gl_code'] != ''].copy()

    meta = os_df[['gl_code', 'os_leaf_code', 'os_description']].drop_duplicates('gl_code').merge(
        bfc[['gl_code', 'sap_mapping', 'gl_name', 'sap_description']].drop_duplicates('gl_code'),
        on='gl_code', how='outer'
    )
    meta['line_items'] = meta['os_leaf_code'].map(lambda x: sorted(ancestors_of(x, nodes)) if isinstance(x, str) and x else [])

    mapping_rows = meta[['gl_code', 'os_leaf_code', 'sap_mapping', 'gl_name', 'sap_description']].fillna('').to_dict(orient='records')
    return meta, mapping_rows


def _read_excel_all_sheets(file_storage) -> dict[str, pd.DataFrame]:
    stream = getattr(file_storage, 'stream', file_storage)
    stream.seek(0)
    return pd.read_excel(stream, sheet_name=None, header=None)


def _find_header_row(df: pd.DataFrame, required: list[str], aliases: dict[str, list[str]] | None = None, scan_rows: int = 20) -> tuple[int, list[str]]:
    aliases = aliases or {}
    required_sets = {name: {name, *aliases.get(name, [])} for name in required}
    limit = min(scan_rows, len(df))
    for i in range(limit):
        row_vals = [str(x).strip() if pd.notna(x) else '' for x in df.iloc[i].tolist()]
        row_set = set(row_vals)
        if all(any(alias in row_set for alias in required_sets[name]) for name in required):
            return i, row_vals
    raise ValueError(f"Could not find a header row containing: {', '.join(required)}")


def _canonicalize_columns(df: pd.DataFrame, aliases: dict[str, list[str]]) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        c = str(col).strip()
        for canonical, opts in aliases.items():
            if c == canonical or c in opts:
                rename[col] = canonical
                break
    return df.rename(columns=rename)


def read_sap(file_storage) -> pd.DataFrame:
    required = ['Month', 'P&L Nos', 'P&L Head', 'GL Code', 'GL Name', 'CC Code', 'Cost Center Name', 'MIS Type', 'SAP Mapping', 'SAP Description', 'Amount']
    aliases = {
        'GL Code': ['GL code', 'GLCode'],
        'GL Name': ['GL name', 'GLName'],
        'SAP Mapping': ['SAP mapping', 'SAPMapping'],
        'SAP Description': ['SAP description', 'SAPDescription'],
    }
    sheets = _read_excel_all_sheets(file_storage)
    picked = None
    for sheet_name, raw in sheets.items():
        try:
            header_idx, header = _find_header_row(raw, ['GL Code', 'SAP Mapping', 'Amount'], aliases=aliases)
            picked = (sheet_name, raw, header_idx, header)
            break
        except Exception:
            continue
    if picked is None:
        raise ValueError('Unable to find a valid header row in the ERP to SAP BFC workbook.')
    _, raw, header_idx, header = picked
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = header
    df = _canonicalize_columns(df, aliases)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ERP to SAP BFC workbook is missing required columns: {', '.join(missing)}")
    df = df[required].copy()
    df = df[~df.apply(lambda r: all(v is None or str(v).strip() == '' for v in r), axis=1)].copy()
    return df.reset_index(drop=True)


def read_os(file_storage) -> pd.DataFrame:
    sheets = _read_excel_all_sheets(file_storage)
    aliases = {
        'OS COA': ['OS Account', 'OS account', 'OS COA '],
        'Local COA': ['GL Code', 'LocalCOA', 'Local COA '],
        'Function': ['P&L Head', 'Function '],
        'Amount': ['Amount ', 'Amt'],
    }
    picked = None
    for _, raw in sheets.items():
        try:
            header_idx, header = _find_header_row(raw, ['Local COA', 'Amount'], aliases=aliases)
            picked = (raw, header_idx, header)
            break
        except Exception:
            continue
    if picked is None:
        raise ValueError('Unable to find a valid header row in the ERP to OS workbook.')
    raw, header_idx, header = picked
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = header
    df = _canonicalize_columns(df, aliases)
    df.columns = [str(c).strip() for c in df.columns]
    if 'OS COA' not in df.columns and 'OS Account' in df.columns:
        df = df.rename(columns={'OS Account': 'OS COA'})
    if 'Function' not in df.columns:
        # derive a broad function label from OS COA or leave blank; V7 line-item membership comes from embedded mapping workbook
        df['Function'] = ''
    required = ['OS COA', 'Amount', 'Local COA', 'Function']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ERP to OS workbook is missing required columns: {', '.join(missing)}")
    df = df[~df.apply(lambda r: all(v is None or str(v).strip() == '' for v in r), axis=1)].copy()
    return df.reset_index(drop=True)


def build_gl_totals(sap: pd.DataFrame, os_df: pd.DataFrame, meta: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    sap = sap.copy()
    sap['gl_code_raw'] = sap['GL Code'].map(extract_gl_code)
    sap['sap_mapping_clean'] = sap['SAP Mapping'].map(normalize_code)
    sap['amount_display'] = -pd.to_numeric(sap['Amount'], errors='coerce').fillna(0.0) / 1000.0

    sap_map_lookup = meta[['gl_code', 'sap_mapping']].dropna().copy()
    sap_map_lookup['sap_mapping'] = sap_map_lookup['sap_mapping'].map(normalize_code)
    sap_map_lookup['gl_code'] = sap_map_lookup['gl_code'].map(normalize_code)
    sap_map_lookup = sap_map_lookup[sap_map_lookup['sap_mapping'] != ''].drop_duplicates('sap_mapping')
    sap = sap.merge(
        sap_map_lookup.rename(columns={'gl_code': 'gl_code_from_mapping'}),
        left_on='sap_mapping_clean',
        right_on='sap_mapping',
        how='left'
    )
    sap['gl_code'] = sap['gl_code_raw']
    sap.loc[sap['gl_code'].eq('') & sap['gl_code_from_mapping'].notna(), 'gl_code'] = sap['gl_code_from_mapping']

    sap_grp = sap[sap['gl_code'] != ''].groupby('gl_code', dropna=False).agg(
        sap_bfc=('amount_display', 'sum'),
        sap_description=('GL Name', first_nonblank),
        sap_currency=('Month', lambda s: ''),
    ).reset_index()

    os_df = os_df.copy()
    os_df['gl_code'] = os_df['Local COA'].map(extract_gl_code)
    os_df['amount_display'] = -pd.to_numeric(os_df['Amount'], errors='coerce').fillna(0.0) / 1000.0
    currency_col = next((c for c in ['Currency', 'Curr', 'Local Currency', 'CCY'] if c in os_df.columns), None)
    if currency_col:
        os_df['currency'] = os_df[currency_col].fillna('').astype(str).str.strip()
    else:
        os_df['currency'] = ''
    os_grp = os_df[os_df['gl_code'] != ''].groupby('gl_code', dropna=False).agg(
        onestream=('amount_display', 'sum'),
        os_currency=('currency', first_nonblank),
        function=('Function', first_nonblank),
    ).reset_index()

    all_gl = meta.merge(sap_grp, on='gl_code', how='outer').merge(os_grp, on='gl_code', how='outer')
    all_gl['sap_bfc'] = pd.to_numeric(all_gl.get('sap_bfc'), errors='coerce').fillna(0.0)
    all_gl['onestream'] = pd.to_numeric(all_gl.get('onestream'), errors='coerce').fillna(0.0)
    all_gl['difference'] = all_gl['sap_bfc'] - all_gl['onestream']
    all_gl['description'] = all_gl['gl_name'].where(all_gl.get('gl_name', '').fillna('').astype(str).str.strip() != '', all_gl.get('os_description', '')).fillna('')
    all_gl['currency'] = all_gl.get('os_currency', '').fillna('')
    all_gl['line_items'] = all_gl['line_items'].apply(lambda x: x if isinstance(x, list) else [])
    all_gl['gl_code'] = all_gl['gl_code'].fillna('').astype(str)
    all_gl = all_gl[all_gl['gl_code'].str.strip() != ''].copy()

    debug = {
        'sap_rows': int(len(sap)),
        'os_rows': int(len(os_df)),
        'sap_gl_mapped': int(sap_grp['gl_code'].nunique()),
        'os_gl_mapped': int(os_grp['gl_code'].nunique()),
        'all_gl_codes': int(all_gl['gl_code'].nunique()),
        'os_rows_with_amount': int((pd.to_numeric(os_df['Amount'], errors='coerce').fillna(0.0) != 0).sum()),
        'os_gl_codes_matched_to_mapping': int(all_gl[(all_gl['onestream'] != 0) & (all_gl['line_items'].apply(len) > 0)]['gl_code'].nunique()),
        'os_gl_codes_unmapped': int(all_gl[(all_gl['onestream'] != 0) & (all_gl['line_items'].apply(len) == 0)]['gl_code'].nunique()),
    }
    return all_gl, debug


def line_item_gls(all_gl: pd.DataFrame, line_code: str) -> pd.DataFrame:
    return all_gl[all_gl['line_items'].apply(lambda xs: line_code in xs)].copy()


def build_summary_rows(all_gl: pd.DataFrame) -> list[dict[str, Any]]:
    rows = []
    for code, name in SUMMARY_ROWS:
        part = line_item_gls(all_gl, code)
        sap = round(float(part['sap_bfc'].sum()), 2)
        os = round(float(part['onestream'].sum()), 2)
        rows.append({
            'code': code,
            'name': name,
            'sap_bfc': sap,
            'onestream': os,
            'difference': round(sap - os, 2),
            'highlight': code in HIGHLIGHT_CODES,
        })
    return rows


def build_drilldown_rows(all_gl: pd.DataFrame) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for code, name in DRILLDOWN_ROWS:
        part = line_item_gls(all_gl, code).sort_values(['difference', 'sap_bfc', 'onestream'], key=lambda s: s.abs(), ascending=False)
        sap = round(float(part['sap_bfc'].sum()), 2)
        os = round(float(part['onestream'].sum()), 2)
        out.append({
            'row_type': 'parent',
            'code': code,
            'name': name,
            'description': '',
            'currency': '',
            'sap_bfc': sap,
            'onestream': os,
            'difference': round(sap - os, 2),
            'highlight': code in HIGHLIGHT_CODES,
            'child_count': int(len(part)),
        })
        for _, r in part.iterrows():
            out.append({
                'row_type': 'child',
                'parent_code': code,
                'code': r['gl_code'],
                'name': r['gl_code'],
                'description': r.get('description', '') or '',
                'currency': r.get('currency', '') or '',
                'sap_bfc': round(float(r['sap_bfc']), 2),
                'onestream': round(float(r['onestream']), 2),
                'difference': round(float(r['difference']), 2),
                'highlight': False,
            })
    return out


def process_files(sap_file, os_file) -> dict[str, Any]:
    nodes = parse_hierarchy(HIERARCHY_PATH)
    meta, mapping_rows = load_v7_mapping(MAPPING_WORKBOOK_PATH, nodes)
    sap = read_sap(sap_file)
    os_df = read_os(os_file)
    all_gl, debug = build_gl_totals(sap, os_df, meta)
    summary_rows = build_summary_rows(all_gl)
    drilldown_rows = build_drilldown_rows(all_gl)

    unmapped_gls = all_gl[all_gl['line_items'].apply(len) == 0].sort_values('difference', key=lambda s: s.abs(), ascending=False).head(200)
    unmapped_rows = [{
        'gl_code': r['gl_code'],
        'description': r.get('description', '') or '',
        'sap_bfc': round(float(r['sap_bfc']), 2),
        'onestream': round(float(r['onestream']), 2),
        'difference': round(float(r['difference']), 2),
    } for _, r in unmapped_gls.iterrows()]

    debug.update({
        'sap_total_all_rows': round(float(all_gl['sap_bfc'].sum()), 2),
        'os_total_all_rows': round(float(all_gl['onestream'].sum()), 2),
        'unmapped_gl_codes': int((all_gl['line_items'].apply(len) == 0).sum()),
        'unmapped_top_items': unmapped_rows,
    })

    return {
        'summary_rows': summary_rows,
        'drilldown_rows': drilldown_rows,
        'debug': debug,
        'mapping_rows': mapping_rows[:1000],
    }


def build_export_workbook(results: dict[str, Any]) -> bytes:
    output = io.BytesIO()
    summary_df = pd.DataFrame(results['summary_rows'])
    drill_df = pd.DataFrame(results['drilldown_rows'])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        drill_df.to_excel(writer, index=False, sheet_name='Drilldown')
        drill_df[(drill_df['row_type'] == 'parent') | (drill_df['sap_bfc'] != 0) | (drill_df['onestream'] != 0)].to_excel(writer, index=False, sheet_name='Drilldown(no zeroes)')
        wb = writer.book
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    value = '' if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                    if ws.title.startswith('Drilldown') and cell.row > 1:
                        code_value = ws.cell(cell.row, 2).value
                        if code_value in HIGHLIGHT_CODES:
                            cell.font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)
    output.seek(0)
    return output.getvalue()


@app.get('/')
def index():
    return render_template('index.html')


@app.post('/api/run-recon')
def run_recon():
    sap_file = request.files.get('sap_file')
    os_file = request.files.get('os_file')
    if not sap_file or not os_file:
        return jsonify({'error': 'Please upload both SAP BFC and OneStream files.'}), 400
    try:
        return jsonify(process_files(sap_file, os_file))
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


@app.post('/api/export')
def export_results():
    sap_file = request.files.get('sap_file')
    os_file = request.files.get('os_file')
    if not sap_file or not os_file:
        return jsonify({'error': 'Please upload both SAP BFC and OneStream files.'}), 400
    try:
        content = build_export_workbook(process_files(sap_file, os_file))
        return send_file(
            io.BytesIO(content),
            as_attachment=True,
            download_name='reconciliation_output_v9.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
